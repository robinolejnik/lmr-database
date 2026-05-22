"""Inspect every sheet in a.xlsx + b.xlsx and produce a column report.

For each column we record:
- non_null count, null count, sampled count
- distinct values (capped) and their frequencies (for low-cardinality cols)
- type observations: how many values look like uuid / bool / int / float / datetime / text
- min/max for numeric columns
- min/max length for text columns
- a guess at the SQL type

Output: schema_report.json (per-sheet, per-column) + tables.json (sheet meta).

Big sheets are sampled (first N rows) — enough for type inference; junction/
small tables are scanned fully.
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

FILES = ["a.xlsx", "b.xlsx"]
SKIP_SHEETS = {"Tabelle1", "Tabelle9"}  # empty placeholder sheets

# Per-sheet row sample cap. We use the full sheet if smaller.
SAMPLE_CAP = 8000
DISTINCT_CAP = 60  # remember up to this many distinct values per col
LOW_CARD_THRESHOLD = 30  # cols with <= this many distincts get full distinct dump

UUID_RE = re.compile(r"^[0-9A-Fa-f]{8}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{12}$")
# Leading-zero strings ("01", "00") are *codes*, not integers — keep as text.
INT_RE = re.compile(r"^-?(0|[1-9]\d*)$")
FLOAT_RE = re.compile(r"^-?\d+\.\d+$")


@dataclass
class ColStats:
    name: str
    sampled: int = 0
    non_null: int = 0
    n_uuid: int = 0
    n_bool: int = 0
    n_int: int = 0
    n_float: int = 0
    n_datetime: int = 0
    n_text: int = 0
    min_num: float | None = None
    max_num: float | None = None
    min_len: int | None = None
    max_len: int | None = None
    distincts: Counter = field(default_factory=Counter)
    distinct_overflow: bool = False

    def add(self, v):
        self.sampled += 1
        if v is None or (isinstance(v, str) and v.strip() == ""):
            return
        self.non_null += 1
        # remember distinct values up to cap
        if len(self.distincts) < DISTINCT_CAP or v in self.distincts:
            self.distincts[v] += 1
        else:
            self.distinct_overflow = True

        if isinstance(v, bool):  # must come before int
            self.n_bool += 1
            return
        if isinstance(v, (datetime, date)):
            self.n_datetime += 1
            return
        if isinstance(v, int):
            self.n_int += 1
            self._upd_num(float(v))
            return
        if isinstance(v, float):
            if v.is_integer():
                self.n_int += 1
            else:
                self.n_float += 1
            self._upd_num(float(v))
            return
        if isinstance(v, str):
            s = v.strip()
            if UUID_RE.match(s):
                self.n_uuid += 1
                return
            if s.lower() in ("true", "false"):
                self.n_bool += 1
                return
            if INT_RE.match(s):
                self.n_int += 1
                self._upd_num(float(s))
                return
            if FLOAT_RE.match(s):
                self.n_float += 1
                self._upd_num(float(s))
                return
            # fall through to text
            self.n_text += 1
            ln = len(s)
            if self.min_len is None or ln < self.min_len:
                self.min_len = ln
            if self.max_len is None or ln > self.max_len:
                self.max_len = ln
            return
        # unknown — treat as text
        self.n_text += 1

    def _upd_num(self, x: float):
        if self.min_num is None or x < self.min_num:
            self.min_num = x
        if self.max_num is None or x > self.max_num:
            self.max_num = x

    def guess_type(self) -> str:
        """Return a best-guess SQL type or descriptor."""
        if self.non_null == 0:
            return "EMPTY"
        # >= 99% in a category wins
        cats = {
            "uuid": self.n_uuid,
            "bool": self.n_bool,
            "datetime": self.n_datetime,
            "int": self.n_int,
            "float": self.n_float,
            "text": self.n_text,
        }
        dominant, n = max(cats.items(), key=lambda kv: kv[1])
        if n / self.non_null < 0.95:
            return f"MIXED({cats})"
        if dominant == "uuid":
            return "uuid"
        if dominant == "bool":
            return "boolean"
        if dominant == "datetime":
            return "timestamp"
        if dominant == "int":
            # Always integer — sampling can't prove smallint is safe.
            # The savings are marginal vs the risk of being burned by values
            # outside the sample window (the way we were with smallint).
            if self.min_num is not None and -2147483648 <= self.min_num and self.max_num <= 2147483647:
                return "integer"
            return "bigint"
        if dominant == "float":
            return "numeric"
        if dominant == "text":
            return "text"
        return "unknown"

    def to_dict(self):
        # serialize distincts: cap, convert datetimes to iso
        def _ser(x):
            if isinstance(x, (datetime, date)):
                return x.isoformat()
            return x

        distincts_list = [
            [_ser(k), c] for k, c in self.distincts.most_common(DISTINCT_CAP)
        ]
        n_distinct = len(self.distincts) + (1 if self.distinct_overflow else 0)
        return {
            "name": self.name,
            "sampled": self.sampled,
            "non_null": self.non_null,
            "null_pct": round(100 * (1 - self.non_null / max(self.sampled, 1)), 2),
            "type_guess": self.guess_type(),
            "type_counts": {
                "uuid": self.n_uuid,
                "bool": self.n_bool,
                "int": self.n_int,
                "float": self.n_float,
                "datetime": self.n_datetime,
                "text": self.n_text,
            },
            "n_distinct_seen": len(self.distincts),
            "distinct_overflow": self.distinct_overflow,
            "min_num": self.min_num,
            "max_num": self.max_num,
            "min_len": self.min_len,
            "max_len": self.max_len,
            "distincts": distincts_list if n_distinct <= LOW_CARD_THRESHOLD else distincts_list[:10],
        }


def analyze_sheet(wb, name: str) -> dict:
    ws = wb[name]
    print(f"  scanning {name} ({ws.max_row} rows × {ws.max_column} cols)...", flush=True)
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    cols = [ColStats(name=h or f"col_{i}") for i, h in enumerate(header)]

    sample_limit = min(SAMPLE_CAP, ws.max_row - 1)
    for i, row in enumerate(rows_iter):
        if i >= sample_limit:
            break
        for c, v in zip(cols, row):
            c.add(v)

    return {
        "sheet": name,
        "n_rows_total": ws.max_row - 1,
        "n_cols": ws.max_column,
        "sampled_rows": sample_limit,
        "columns": [c.to_dict() for c in cols],
    }


def main():
    # repo layout: <root>/tools/migration/analyze.py — root is two parents up
    here = Path(__file__).resolve().parent
    repo_root = here.parent.parent
    data_dir = repo_root / "data"
    report = {"sheets": {}}
    for fname in FILES:
        path = data_dir / fname
        print(f"\n=== {fname} ===", flush=True)
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            if sheet in SKIP_SHEETS:
                continue
            try:
                report["sheets"][sheet] = analyze_sheet(wb, sheet)
                report["sheets"][sheet]["source_file"] = fname
            except Exception as e:
                print(f"    !! failed: {e}", flush=True)
                report["sheets"][sheet] = {"error": str(e), "source_file": fname}
        wb.close()

    out = here / "schema_report.json"
    out.write_text(json.dumps(report, indent=2, default=str))
    print(f"\nwrote {out}")


if __name__ == "__main__":
    main()
